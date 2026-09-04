# -*- coding: utf-8 -*-
"""Genere un classeur par ticket, a la charte, pret a etre televerse sur le Drive.

Forme reprise des extractions existantes du client : ligne 1 = libelles de
colonnes sur fond #1A73E8 en blanc, ligne 2 = bandeau pedagogique
« Important ⚠️ : … » fusionne sur toute la largeur, donnees a partir de la
ligne 3. Comfortaa 10, centrage horizontal et vertical, deux premieres lignes
figees, aucune colonne vide.
"""
import os, sys, re
ICI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ICI)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLEU, POLICE = "1A73E8", "Comfortaa"
BORD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
SORTIE = os.path.join(ICI, "sheets")


def charger():
    fichiers = {}
    for mod in ("contenu_sheets_1", "contenu_sheets_2", "contenu_sheets_3", "contenu_sheets_4"):
        fichiers.update(__import__(mod).FICHIERS)
    return fichiers


def largeur(colonne_valeurs, entete):
    """Largeur lisible sans etre demesuree."""
    longueurs = [len(str(v)) for v in colonne_valeurs if v] + [len(entete)]
    return max(14, min(58, int(max(longueurs) * 0.95) + 2))


def construire(nom, bandeau, entetes, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[^A-Z0-9 ]", "", nom.upper())[:31].strip()
    n = len(entetes)

    for i, libelle in enumerate(entetes, start=1):
        c = ws.cell(row=1, column=i, value=libelle)
        c.font = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORD
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    b = ws.cell(row=2, column=1, value=bandeau)
    b.font = Font(name=POLICE, size=10, italic=True)
    b.fill = PatternFill("solid", fgColor="E8F0FE")
    b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(1, n + 1):
        ws.cell(row=2, column=i).border = BORD
    ws.row_dimensions[2].height = 62
    ws.freeze_panes = "A3"

    for r, ligne in enumerate(lignes, start=3):
        # une ligne de sous-titre (2 cellules remplies sur une grille large) reste lisible
        for i in range(1, n + 1):
            val = ligne[i - 1] if i - 1 < len(ligne) else ""
            c = ws.cell(row=r, column=i, value=val if val != "" else None)
            c.font = Font(name=POLICE, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORD
        ws.row_dimensions[r].height = 40

    for i in range(1, n + 1):
        colonne = [l[i - 1] if i - 1 < len(l) else "" for l in lignes]
        ws.column_dimensions[get_column_letter(i)].width = largeur(colonne, entetes[i - 1])

    ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"
    chemin = os.path.join(SORTIE, nom + ".xlsx")
    wb.save(chemin)
    return chemin


if __name__ == "__main__":
    os.makedirs(SORTIE, exist_ok=True)
    for nom, (bandeau, entetes, lignes) in charger().items():
        p = construire(nom, bandeau, entetes, lignes)
        print(f"{os.path.getsize(p):>7} o  {os.path.basename(p)}")
