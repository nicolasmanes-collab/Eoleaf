# -*- coding: utf-8 -*-
"""Construit le classeur de roadmap technique + les TSV prêts à coller.

Charte appliquée (décrite dans le skill audit-technique, faute d'accès au
standard.json partagé) : Comfortaa 10, centrage horizontal et vertical,
libellés de colonnes en ligne 1 sur fond #1A73E8 en blanc, bandeau
pédagogique fusionné en ligne 2, les deux figés, feuilles en MAJUSCULES.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from donnees_roadmap import SEPTEMBRE, OCTOBRE, NOVEMBRE, BACKLOG, CONSTATS

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLEU, POLICE = "1A73E8", "Comfortaa"
BORD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
TEINTE = {"bloque": "FBE9E7", "freine": "FFF8E1", "crawl": "E8F0FE"}

COLS = ["MOIS", "DATE DE LIVRAISON", "TICKET", "DETAILS", "REPARTITION",
        "HEURE DE TRAVAIL", "STATUT", "LIVRABLE", "INTERVENANT"]
LARG = [13, 16, 32, 104, 20, 12, 12, 30, 17]


def _entetes(ws, cols):
    for i, nom in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=nom)
        c.font = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORD
    ws.row_dimensions[1].height = 32


def _bandeau(ws, ncol, texte):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(row=2, column=1, value=texte)
    c.font = Font(name=POLICE, size=10, italic=True)
    c.fill = PatternFill("solid", fgColor="E8F0FE")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(1, ncol + 1):
        ws.cell(row=2, column=i).border = BORD
    ws.row_dimensions[2].height = 58
    ws.freeze_panes = "A3"


def feuille_tickets(wb, titre, lots, bandeau):
    ws = wb.create_sheet(titre)
    _entetes(ws, COLS)
    _bandeau(ws, len(COLS), bandeau)
    r = 3
    for lignes in lots:
        for ligne in lignes:
            severite = ligne[9]
            for i in range(1, len(COLS) + 1):
                val = ligne[i - 1]
                c = ws.cell(row=r, column=i)
                if i == 8 and val:
                    libelle, _, url = str(val).partition("|")
                    c.value = libelle
                    c.hyperlink = url
                    c.font = Font(name=POLICE, size=10, color="1155CC", underline="single")
                else:
                    c.value = val if val != "" else None
                    c.font = Font(name=POLICE, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = BORD
                if i == 1:
                    c.fill = PatternFill("solid", fgColor=TEINTE[severite])
            ws.row_dimensions[r].height = 104
            r += 1
        # total du lot, hors zone de donnees
        ws.cell(row=r, column=5, value="TOTAL " + lignes[0][0]).font = Font(name=POLICE, size=10, bold=True)
        t = ws.cell(row=r, column=6, value=sum(l[5] for l in lignes))
        t.font = Font(name=POLICE, size=10, bold=True)
        for col in (5, 6):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    for i, l in enumerate(LARG, start=1):
        ws.column_dimensions[get_column_letter(i)].width = l
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    return ws


def feuille_constats(wb):
    ws = wb.create_sheet("CONSTATS CHIFFRES")
    cols = ["VOLET", "CE QUI EST MESURE", "VALEUR RELEVEE", "CE QUE CELA IMPLIQUE"]
    _entetes(ws, cols)
    _bandeau(ws, len(cols),
             "Important ⚠️ : chaque ticket de la roadmap s'appuie sur une de ces valeurs. Toutes viennent "
             "des extractions du Drive (crawl du 20/05/2026), du relevé H1 du 24/08/2026 ou de l'onglet "
             "sémantique. Un constat sans chiffre daté ne se compare pas d'un mois sur l'autre.")
    for r, ligne in enumerate(CONSTATS, start=3):
        for i, val in enumerate(ligne, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = Font(name=POLICE, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORD
        ws.row_dimensions[r].height = 44
    for i, l in enumerate([16, 46, 22, 80], start=1):
        ws.column_dimensions[get_column_letter(i)].width = l
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws


def main():
    out = os.path.dirname(os.path.abspath(__file__))
    wb = Workbook()
    wb.remove(wb.active)

    feuille_tickets(
        wb, "ROADMAP TECHNIQUE", [SEPTEMBRE],
        "Important ⚠️ : l'ordre des lignes est l'ordre d'impact sur le chiffre d'affaires, pas la sévérité "
        "du crawler. La teinte de la colonne MOIS le dit : rouge, la page ne peut pas vendre ; ambre, elle "
        "vend moins bien qu'elle ne devrait ; bleu, du budget de crawl part en fumée. Chaque ligne porte son "
        "critère d'acceptation dans DETAILS et son Sheet de travail dans LIVRABLE.")

    feuille_tickets(
        wb, "BACKLOG OCT-NOV", [OCTOBRE, NOVEMBRE],
        "Important ⚠️ : ces chantiers sont identifiés et outillés (leur Sheet de travail existe déjà) mais "
        "ne tiennent pas dans le volume de septembre. Quatre volets du process n'ont jamais été ouverts "
        "depuis le début de l'accompagnement : hreflang, vitesse, JSON-LD et maillage interne.")

    feuille_constats(wb)

    chemin = os.path.join(out, "2026-09-04-roadmap-technique-septembre-eoleaf.xlsx")
    wb.save(chemin)
    print("classeur :", chemin)

    # TSV au format exact de l'onglet technique du Sheet client
    tsv = os.path.join(out, "a-coller-onglet-technique.tsv")
    with open(tsv, "w", encoding="utf-8") as f:
        f.write("Mois\tSemaine\tBrief technique\tLivrable\tIntervenant\tStatut\n")
        for l in SEPTEMBRE + BACKLOG:
            if l[7]:
                libelle, _, url = l[7].partition("|")
                livrable = f'=LIEN_HYPERTEXTE("{url}";"{libelle}")'
            else:
                livrable = ""
            f.write(f"{l[0]}\t{l[1]}\t{l[2]}\t{livrable}\t{l[8]}\tA faire\n")
    print("tsv sheet :", tsv)

    # TSV complet au format du modele de roadmap technique
    tsv2 = os.path.join(out, "roadmap-technique-eoleaf-complet.tsv")
    with open(tsv2, "w", encoding="utf-8") as f:
        f.write("\t".join(COLS + ["SEVERITE"]) + "\n")
        for l in SEPTEMBRE + BACKLOG:
            f.write("\t".join(str(x).replace("\t", " ") for x in l) + "\n")
    print("tsv complet :", tsv2)


if __name__ == "__main__":
    main()
